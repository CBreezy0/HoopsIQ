from __future__ import annotations

import json
import logging
from pathlib import Path
import subprocess
import sys
import warnings


warnings.filterwarnings("ignore", message="urllib3 v2 only supports OpenSSL")


REPO_ROOT = Path(__file__).resolve().parents[1]
APP_DIR = REPO_ROOT / "app"
CONFIG_DIR = REPO_ROOT / "config"
DATA_DIR = REPO_ROOT / "data"
OUTPUTS_DIR = REPO_ROOT / "outputs"
LOGS_DIR = REPO_ROOT / "logs"
DEFAULT_CONFIG_PATH = CONFIG_DIR / "config.json"

_CONFIG_CACHE: dict[Path, dict] = {}
RATINGS_OUTPUT_PATH = OUTPUTS_DIR / "teams_power_full.xlsx"
GAMES_HISTORY_PATH = OUTPUTS_DIR / "games_history.csv"
TEAM_DIRECTORY_NAMES_PATH = OUTPUTS_DIR / "team_directory_names.json"
TEAM_NAME_MAP_PATH = OUTPUTS_DIR / "team_name_map.json"
TEAM_ID_MAP_PATH = OUTPUTS_DIR / "team_id_map.json"
TEAM_ALIASES_PATH = DATA_DIR / "team_aliases.json"
LIVE_REQUIRED_OUTPUTS = (
    OUTPUTS_DIR / "game_predictions.csv",
    OUTPUTS_DIR / "predictions_log.csv",
)


def resolve_project_root(config_path: str | Path | None = None) -> Path:
    if config_path is None:
        return REPO_ROOT
    path = Path(config_path).expanduser().resolve()
    if path.parent == CONFIG_DIR:
        return REPO_ROOT
    return path.parent


def load_public_config(config_path: str | Path | None = None, *, refresh: bool = False) -> dict:
    path = Path(config_path or DEFAULT_CONFIG_PATH).expanduser().resolve()
    if not refresh and path in _CONFIG_CACHE:
        return _CONFIG_CACHE[path]
    if not path.exists():
        _CONFIG_CACHE[path] = {}
        return _CONFIG_CACHE[path]
    with path.open("r", encoding="utf-8") as fh:
        payload = json.load(fh)
    _CONFIG_CACHE[path] = payload if isinstance(payload, dict) else {}
    return _CONFIG_CACHE[path]


def ensure_runtime_dirs() -> None:
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)


def build_public_logger(name: str, *, debug: bool = False) -> logging.Logger:
    cfg = load_public_config()
    logging_cfg = cfg.get("logging", {}) if isinstance(cfg.get("logging", {}), dict) else {}
    level_name = str(
        logging_cfg.get("debug_level" if debug else "level", "INFO")
    ).upper()
    format_key = "debug_format" if debug else "format"
    fmt = str(logging_cfg.get(format_key, "%(asctime)s %(levelname)s %(message)s"))
    level = getattr(logging, level_name, logging.INFO)

    logger = logging.getLogger(name)
    logger.handlers.clear()
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter(fmt))
    logger.addHandler(handler)
    logger.setLevel(level)
    logger.propagate = False
    return logger


def _csv_has_data_rows(path: Path) -> bool:
    if not path.exists() or not path.is_file():
        return False
    try:
        if path.stat().st_size <= 0:
            return False
    except Exception:
        return False

    try:
        with path.open("r", encoding="utf-8", errors="ignore") as fh:
            header_found = False
            for raw_line in fh:
                if raw_line.strip():
                    header_found = True
                    break
            if not header_found:
                return False
            for raw_line in fh:
                if raw_line.strip():
                    return True
    except Exception:
        return False
    return False


def _csv_row_count(path: Path) -> int | None:
    if not path.exists() or not path.is_file():
        return None
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as fh:
            saw_header = False
            rows = 0
            for raw_line in fh:
                if not raw_line.strip():
                    continue
                if not saw_header:
                    saw_header = True
                    continue
                rows += 1
            return rows if saw_header else 0
    except Exception:
        return None


def _json_entry_count(path: Path) -> int | None:
    if not path.exists() or not path.is_file():
        return None
    try:
        with path.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
        if isinstance(payload, dict):
            return len(payload)
        if isinstance(payload, list):
            return len(payload)
    except Exception:
        return None
    return None


def _xlsx_row_count(path: Path) -> int | None:
    if not path.exists() or not path.is_file():
        return None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            max_row = int(ws.max_row or 0)
            return max(max_row - 1, 0) if max_row > 0 else 0
        finally:
            wb.close()
    except Exception:
        return None


def _file_exists_with_content(path: Path) -> bool:
    if not path.exists() or not path.is_file():
        return False
    try:
        return path.stat().st_size > 0
    except Exception:
        return False


def ratings_output_status() -> dict[str, object]:
    row_count = _xlsx_row_count(RATINGS_OUTPUT_PATH)
    ready = bool(_file_exists_with_content(RATINGS_OUTPUT_PATH) and (row_count or 0) > 0)
    return {
        "ready": ready,
        "row_count": row_count,
        "missing_or_empty": [] if ready else [str(RATINGS_OUTPUT_PATH.relative_to(REPO_ROOT))],
    }


def lookup_data_status() -> dict[str, object]:
    games_rows = _csv_row_count(GAMES_HISTORY_PATH)
    team_name_rows = _json_entry_count(TEAM_NAME_MAP_PATH)
    team_dir_name_rows = _json_entry_count(TEAM_DIRECTORY_NAMES_PATH)
    team_id_rows = _json_entry_count(TEAM_ID_MAP_PATH)
    aliases_rows = _json_entry_count(TEAM_ALIASES_PATH)

    missing_or_empty: list[str] = []
    if not (games_rows and games_rows > 0):
        missing_or_empty.append(str(GAMES_HISTORY_PATH.relative_to(REPO_ROOT)))
    if max(team_name_rows or 0, team_dir_name_rows or 0, team_id_rows or 0) <= 0:
        missing_or_empty.append(str(TEAM_DIRECTORY_NAMES_PATH.relative_to(REPO_ROOT)))

    return {
        "ready": not missing_or_empty,
        "games_rows": games_rows,
        "team_name_rows": max(team_name_rows or 0, team_dir_name_rows or 0),
        "team_id_rows": team_id_rows,
        "aliases_rows": aliases_rows,
        "missing_or_empty": missing_or_empty,
    }


def live_outputs_status() -> dict[str, object]:
    missing_or_empty = [
        str(path.relative_to(REPO_ROOT))
        for path in LIVE_REQUIRED_OUTPUTS
        if not _csv_has_data_rows(path)
    ]
    return {
        "ready": not missing_or_empty,
        "missing_or_empty": missing_or_empty,
    }


def log_startup_file_checks(logger: logging.Logger) -> None:
    games_rows = _csv_row_count(GAMES_HISTORY_PATH)
    ratings_rows = _xlsx_row_count(RATINGS_OUTPUT_PATH)
    team_name_rows = _json_entry_count(TEAM_NAME_MAP_PATH)
    team_dir_name_rows = _json_entry_count(TEAM_DIRECTORY_NAMES_PATH)
    team_id_rows = _json_entry_count(TEAM_ID_MAP_PATH)
    alias_rows = _json_entry_count(TEAM_ALIASES_PATH)
    team_name_map_path = TEAM_NAME_MAP_PATH if TEAM_NAME_MAP_PATH.exists() else TEAM_DIRECTORY_NAMES_PATH

    logger.info(
        "FILE_CHECK games_history exists=%s rows=%s path=%s",
        str(GAMES_HISTORY_PATH.exists()).lower(),
        games_rows if games_rows is not None else "n/a",
        GAMES_HISTORY_PATH.relative_to(REPO_ROOT),
    )
    logger.info(
        "FILE_CHECK teams_power exists=%s rows=%s path=%s",
        str(RATINGS_OUTPUT_PATH.exists()).lower(),
        ratings_rows if ratings_rows is not None else "n/a",
        RATINGS_OUTPUT_PATH.relative_to(REPO_ROOT),
    )
    logger.info(
        "FILE_CHECK team_name_map exists=%s rows=%s path=%s",
        str((TEAM_NAME_MAP_PATH.exists() or TEAM_DIRECTORY_NAMES_PATH.exists())).lower(),
        max(team_name_rows or 0, team_dir_name_rows or 0),
        team_name_map_path.relative_to(REPO_ROOT),
    )
    logger.info(
        "FILE_CHECK team_id_map exists=%s rows=%s path=%s",
        str(TEAM_ID_MAP_PATH.exists()).lower(),
        team_id_rows if team_id_rows is not None else "n/a",
        TEAM_ID_MAP_PATH.relative_to(REPO_ROOT),
    )
    logger.info(
        "FILE_CHECK team_aliases exists=%s rows=%s path=%s",
        str(TEAM_ALIASES_PATH.exists()).lower(),
        alias_rows if alias_rows is not None else "n/a",
        TEAM_ALIASES_PATH.relative_to(REPO_ROOT),
    )


def _run_main_command(args: list[str], logger: logging.Logger, log_message: str) -> bool:
    logger.info(log_message)
    try:
        subprocess.run(
            [sys.executable, str(REPO_ROOT / "main.py"), *args],
            cwd=REPO_ROOT,
            check=True,
        )
        return True
    except Exception as ex:
        logger.warning(
            "BOOTSTRAP command failed args=%s error=%s: %s",
            " ".join(args),
            type(ex).__name__,
            ex,
        )
        return False


def bootstrap_live_data_generation(logger: logging.Logger | None = None) -> dict[str, object]:
    ensure_runtime_dirs()
    logger = logger or build_public_logger("bootstrap")
    log_startup_file_checks(logger)
    ratings_status = ratings_output_status()
    lookup_status = lookup_data_status()
    live_status = live_outputs_status()

    if ratings_status["ready"] and lookup_status["ready"] and live_status["ready"]:
        return {
            "ran": False,
            "ready": True,
            "missing_or_empty": [],
            "ratings_ready": True,
            "lookup_ready": True,
        }

    ran = False
    if not ratings_status["ready"] or not lookup_status["ready"]:
        ran = True
        _run_main_command(
            ["--backfill"],
            logger,
            "BOOTSTRAP building ratings from scratch",
        )
        log_startup_file_checks(logger)
        ratings_status = ratings_output_status()
        lookup_status = lookup_data_status()
        live_status = live_outputs_status()

    if ratings_status["ready"] and lookup_status["ready"] and not live_status["ready"]:
        ran = True
        _run_main_command(
            ["--live"],
            logger,
            "BOOTSTRAP running live predictions",
        )
        log_startup_file_checks(logger)
        live_status = live_outputs_status()

    return {
        "ran": ran,
        "ready": bool(ratings_status["ready"] and lookup_status["ready"] and live_status["ready"]),
        "missing_or_empty": list(lookup_status["missing_or_empty"]) + list(live_status["missing_or_empty"]),
        "ratings_ready": bool(ratings_status["ready"]),
        "lookup_ready": bool(lookup_status["ready"]),
        "ratings_missing_or_empty": list(ratings_status["missing_or_empty"]),
        "lookup_missing_or_empty": list(lookup_status["missing_or_empty"]),
    }
